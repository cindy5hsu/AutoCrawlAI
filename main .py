import os 
import re
import time
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from datetime import datetime
from collections import defaultdict

# 設定基本 URL
BASE_URL = "https://aipure.ai/tw"

# 設定輸出 Excel 檔案
OUTPUT_DIR = r"C:\\detail"
os.makedirs(OUTPUT_DIR, exist_ok=True)

def save_to_excel(file_path, data):
    """將數據存入 Excel"""
    if os.path.exists(file_path):
        workbook = load_workbook(file_path)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["大類別", "小類別數量", "小類別", "工具數量", "工具總數量"])  # 添加標題行
    
    for row in data:
        sheet.append(row)

    # 設置大類別工具總數的格式（粗體 + 明顯顏色）
    for i, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row), start=2):
        if row[4].value:  # 確保這行有大類別工具總數
            for cell in row:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # 黃色背景
    
    workbook.save(file_path)

def fetch_page_selenium(url, wait_time=3):
    """使用 Selenium 獲取動態載入的網頁內容"""
    options = Options()
    options.headless = False  # 看到 Selenium 操作過程
    driver = webdriver.Chrome(options=options)
    driver.get(url)
    print(f"開始爬取: {url}")
    
    time.sleep(wait_time)  # 確保所有內容載入
    page_source = driver.page_source
    driver.quit()
    
    return page_source

# 目標網址
CATEGORY_URL = f"{BASE_URL}/category"
page_source = fetch_page_selenium(CATEGORY_URL)
soup = BeautifulSoup(page_source, "html.parser")

# 解析大類別與小類別
categories = []
category_tool_counts = defaultdict(int) #大類別總數
current_main_category = ""

for div in soup.find_all("div", class_="self-stretch text-base font-bold leading-normal text-slate-600"):
    current_main_category = div.text.strip()
    parent_div = div.find_parent("div")
    if parent_div:
        next_div = parent_div.find_next_sibling("div")
 
        if next_div:
            sub_categories = next_div.find_all("a", class_="flex h-10 items-center gap-2 rounded-lg border border-slate-100 bg-white px-3")
            sub_category_count = len(sub_categories)
            total_tools = 0 # 計數大類別的總工具數
            sub_category_list = []

            for sub_div in sub_categories:
                sub_category = sub_div.text.strip()
                sub_category_link = BASE_URL + sub_div["href"]
                
                sub_page_source = fetch_page_selenium(sub_category_link, wait_time=1)
                sub_soup = BeautifulSoup(sub_page_source, "html.parser")
                tools_count_text = sub_soup.find('div', class_="self-stretch text-sm text-slate-400")
                # tools_count = "未知"  # 預設值
                tools_count = 0  # 預設值
                if tools_count_text:
                    match = re.search(r"探索 (\d+) 個", tools_count_text.get_text(strip=True))
                    if match:
                        tools_count = int(match.group(1))  # 轉換為整數，方便後續計算

                total_tools += tools_count #纍計大類別工具數量
                # categories.append([current_main_category, sub_category_count, sub_category, tools_count])
                sub_category_list.append(["", "", sub_category, tools_count, ""])  # 先填充空白

            # 在小類別列表的最後添加大類別工具總數，並讓它明顯可見
            categories.append([current_main_category, sub_category_count, "", "", total_tools])
            categories.extend(sub_category_list)
            

# **保存到 Excel**
timestamp = datetime.now().strftime("%Y_%m_%d_%H-%M")
excel_file_path = os.path.join(OUTPUT_DIR, f"ai_categories_{timestamp}.xlsx")
save_to_excel(excel_file_path, categories)

print(f"爬取完成，數據已保存到 {excel_file_path}")
